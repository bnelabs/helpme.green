from __future__ import annotations

import csv
import io
import json
import urllib.error
import urllib.parse
import urllib.request
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .domain import CaseFact


class ReadOnlyViolation(PermissionError):
    """An MCP operation attempted to execute or mutate something."""


class _WhitelistRedirectHandler(urllib.request.HTTPRedirectHandler):
    def __init__(self, allowed_hosts: frozenset[str]) -> None:
        self.allowed_hosts = allowed_hosts

    def redirect_request(
        self, req: urllib.request.Request, fp: Any, code: int, msg: str, headers: Any, newurl: str
    ) -> urllib.request.Request | None:
        del fp, code, msg, headers
        parsed = urllib.parse.urlparse(newurl)
        host = (parsed.hostname or "").casefold()
        if parsed.scheme != "https" or host not in self.allowed_hosts:
            raise ReadOnlyViolation(
                "MCP redirect target is not an explicitly whitelisted HTTPS host."
            )
        return urllib.request.Request(
            newurl,
            headers=dict(req.headers),
            origin_req_host=req.origin_req_host,
            unverifiable=True,
            method=req.get_method(),
        )


class ReadOnlyMCP:
    """Narrow read-only import boundary for user files and whitelisted URLs."""

    def __init__(
        self,
        *,
        file_roots: Iterable[Path] = (),
        allowed_url_hosts: Iterable[str] = (),
        max_bytes: int = 2_000_000,
    ) -> None:
        if max_bytes <= 0:
            raise ValueError("MCP read limit must be positive.")
        self.file_roots = tuple(path.resolve() for path in file_roots)
        self.allowed_url_hosts = frozenset(host.casefold() for host in allowed_url_hosts)
        self.max_bytes = max_bytes

    def capabilities(self) -> dict[str, object]:
        """Return a non-secret expert capability description; no paths are exposed."""
        return {
            "read_only": True,
            "tools": ["read_json", "read_csv", "read_xlsx", "fetch_whitelisted_https"],
            "execution": False,
            "writes": False,
            "configured_file_roots": len(self.file_roots),
            "configured_url_hosts": len(self.allowed_url_hosts),
            "max_bytes": self.max_bytes,
        }

    def load(self, target: Path | str) -> tuple[CaseFact, ...]:
        text_target = str(target)
        parsed = urllib.parse.urlparse(text_target)
        if parsed.scheme in {"http", "https"}:
            return self.load_url(text_target)
        return self.load_file(Path(target))

    def load_file(self, path: Path) -> tuple[CaseFact, ...]:
        resolved = path.expanduser().resolve()
        if not any(resolved == root or root in resolved.parents for root in self.file_roots):
            raise ReadOnlyViolation("MCP file access is outside the configured read-only roots.")
        if not resolved.is_file():
            raise FileNotFoundError(str(resolved))
        if resolved.stat().st_size > self.max_bytes:
            raise ReadOnlyViolation("MCP input exceeds the configured read limit.")
        suffix = resolved.suffix.casefold()
        if suffix == ".json":
            return self._facts_from_payload(
                json.loads(resolved.read_text(encoding="utf-8")), str(resolved)
            )
        if suffix == ".csv":
            with resolved.open("r", encoding="utf-8", newline="") as handle:
                rows = list(csv.DictReader(handle))
            return self._facts_from_rows(rows, str(resolved))
        if suffix in {".xlsx", ".xlsm"}:
            workbook = load_workbook(resolved, read_only=True, data_only=True)
            try:
                sheet = workbook.active
                values = list(sheet.iter_rows(values_only=True))
            finally:
                workbook.close()
            if not values:
                return ()
            headings = [str(item or "") for item in values[0]]
            rows = [
                {
                    headings[index]: row[index]
                    for index in range(min(len(headings), len(row)))
                    if headings[index]
                }
                for row in values[1:]
            ]
            return self._facts_from_rows(rows, str(resolved))
        raise ReadOnlyViolation("MCP supports JSON, CSV, and XLSX reads only.")

    def load_url(self, url: str) -> tuple[CaseFact, ...]:
        parsed = urllib.parse.urlparse(url)
        host = (parsed.hostname or "").casefold()
        if parsed.scheme != "https" or host not in self.allowed_url_hosts:
            raise ReadOnlyViolation("MCP URL access requires an explicitly whitelisted HTTPS host.")
        request = urllib.request.Request(url, headers={"User-Agent": "helpme.green/0.1"})
        opener = urllib.request.build_opener(_WhitelistRedirectHandler(self.allowed_url_hosts))
        try:
            with opener.open(request, timeout=20) as response:
                data = response.read(self.max_bytes + 1)
                content_type = response.headers.get_content_type()
        except ReadOnlyViolation:
            raise
        except (OSError, urllib.error.URLError) as exc:
            raise ReadOnlyViolation("MCP URL read failed safely.") from exc
        if len(data) > self.max_bytes:
            raise ReadOnlyViolation("MCP response exceeds the configured read limit.")
        if content_type == "text/csv" or url.casefold().endswith(".csv"):
            rows = list(csv.DictReader(io.StringIO(data.decode("utf-8"))))
            return self._facts_from_rows(rows, url)
        return self._facts_from_payload(json.loads(data.decode("utf-8")), url)

    def execute(self, command: str) -> None:
        del command
        raise ReadOnlyViolation("The MCP contract forbids code execution and mutations.")

    def write(self, target: str, content: str) -> None:
        del target, content
        raise ReadOnlyViolation("The MCP contract forbids writes.")

    def _facts_from_payload(self, payload: Any, reference: str) -> tuple[CaseFact, ...]:
        if isinstance(payload, dict) and isinstance(payload.get("facts"), dict):
            payload = payload["facts"]
        if isinstance(payload, dict):
            return tuple(
                self._fact(str(key), value, reference)
                for key, value in payload.items()
                if str(key) != "_metadata"
            )
        if isinstance(payload, list):
            return self._facts_from_rows(payload, reference)
        raise ReadOnlyViolation("MCP input must be a JSON object or list of rows.")

    def _facts_from_rows(self, rows: list[dict[str, Any]], reference: str) -> tuple[CaseFact, ...]:
        facts: list[CaseFact] = []
        for row in rows:
            for key, value in row.items():
                if key:
                    facts.append(self._fact(str(key), value, reference))
        return tuple(facts)

    @staticmethod
    def _fact(key: str, value: Any, reference: str) -> CaseFact:
        return CaseFact.user(
            key,
            value,
            "unknown" if value is None or value == "" else "declared",
            note="Imported from user-supplied MCP content; unverified and injection-isolated.",
            untrusted=True,
            reference=reference,
        )
