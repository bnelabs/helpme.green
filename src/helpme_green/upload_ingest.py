from __future__ import annotations

import hashlib
import io
import json
import os
import re
import uuid
import zipfile
from collections.abc import Callable, Mapping
from dataclasses import dataclass
from email import policy
from email.parser import BytesParser
from pathlib import Path
from typing import Any

from .knowledge_store import KnowledgeDatabase

MAX_FILES_PER_REQUEST = 10

_ALLOWED_EXTENSIONS = {".pdf", ".html", ".htm", ".xml", ".txt", ".csv", ".json", ".xlsx"}
_EXTENSION_BY_CONTENT_TYPE = {
    "application/pdf": ".pdf",
    "text/html": ".html",
    "application/xhtml+xml": ".html",
    "application/xml": ".xml",
    "text/xml": ".xml",
    "text/plain": ".txt",
    "text/csv": ".csv",
    "application/json": ".json",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
}
_PRINTABLE_THRESHOLD = 0.6
_MAX_PDF_PAGES = 500
_MAX_XLSX_MEMBERS = 1000
_MAX_XLSX_UNCOMPRESSED = 50_000_000
_MAX_XLSX_CELLS = 200_000

_METADATA_FIELDS = {
    "title",
    "publisher",
    "materialFamilies",
    "jurisdiction",
    "authorityTier",
    "scale",
}


class UploadError(ValueError):
    """Raised when an upload fails validation or extraction with a stable error code."""

    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code


@dataclass(frozen=True)
class MultipartPart:
    filename: str
    declared_content_type: str
    data: bytes


@dataclass(frozen=True)
class UploadResult:
    upload_id: str
    filename: str
    status: str
    job_id: str
    duplicate_of: str = ""
    error_code: str = ""
    error_detail: str = ""

    def to_dict(self) -> dict[str, str]:
        return {
            "uploadId": self.upload_id,
            "filename": self.filename,
            "status": self.status,
            "jobId": self.job_id,
            "duplicateOf": self.duplicate_of,
            "errorCode": self.error_code,
            "errorDetail": self.error_detail,
        }


def _normalise_filename(filename: str) -> str:
    # Display-only sanitization; never used for storage keys.
    base = Path(filename.replace("\\", "/")).name
    return re.sub(r"[^\w .-]", "_", base).strip(" .")[:200] or "upload"


def parse_multipart(
    content_type_header: str, body: bytes
) -> tuple[list[MultipartPart], dict[str, str]]:
    """Parse a bounded multipart body with the maintained standard-library email parser."""
    if not content_type_header.casefold().startswith("multipart/form-data"):
        raise UploadError("invalid_content_type", "Request must be multipart/form-data.")
    raw = b"Content-Type: " + content_type_header.encode("latin-1", "replace") + b"\r\n\r\n" + body
    try:
        message = BytesParser(policy=policy.default).parsebytes(raw)
    except Exception as exc:  # the email parser exposes several parse-specific errors.
        raise UploadError("malformed_multipart", "Multipart body could not be parsed.") from exc
    parts: list[MultipartPart] = []
    fields: dict[str, str] = {}
    for part in message.iter_parts():
        filename = part.get_filename()
        payload = part.get_payload(decode=True)
        if not isinstance(payload, bytes):
            payload = b""
        if filename:
            parts.append(
                MultipartPart(
                    filename=_normalise_filename(filename),
                    declared_content_type=str(part.get_content_type() or ""),
                    data=payload,
                )
            )
        else:
            name = part.get_param("name", header="content-disposition")
            if name:
                try:
                    fields[str(name)] = payload.decode("utf-8", errors="replace")[:2000]
                except UnicodeDecodeError:
                    fields[str(name)] = ""
    if not parts:
        raise UploadError("no_files", "Multipart request contained no files.")
    return parts, fields


def _detect_extension(filename: str, data: bytes, declared_content_type: str) -> str:
    filename_extension = Path(filename).suffix.casefold()
    if filename_extension in {".xlsm"}:
        raise UploadError("unsupported_type", "Macro-enabled workbooks are not supported.")
    if filename_extension not in _ALLOWED_EXTENSIONS:
        raise UploadError(
            "unsupported_type", f"File type {filename_extension or 'unknown'} is not supported."
        )
    if data.startswith(b"%PDF-"):
        detected = ".pdf"
    elif data.startswith(b"PK\x03\x04"):
        detected = ".xlsx"
    elif b"<?xml" in data.lstrip()[:200].lower():
        detected = ".xml"
    elif b"<html" in data.lstrip()[:400].lower():
        detected = ".html"
    elif data.lstrip()[:1] in (b"{", b"["):
        detected = ".json"
    else:
        detected = filename_extension
    if detected not in _ALLOWED_EXTENSIONS:
        raise UploadError("unsupported_type", f"Detected file type {detected} is not supported.")
    declared_type = declared_content_type.casefold().split(";", 1)[0].strip()
    if declared_type in _EXTENSION_BY_CONTENT_TYPE:
        declared_extension = _EXTENSION_BY_CONTENT_TYPE[declared_type]
        if detected != declared_extension and detected != filename_extension:
            if declared_extension == ".xlsx" and not data.startswith(b"PK\x03\x04"):
                raise UploadError("mismatched_content_type", "Declared XLSX is not a ZIP workbook.")
            if declared_extension == ".pdf" and not data.startswith(b"%PDF-"):
                raise UploadError("mismatched_content_type", "Declared PDF has no PDF header.")
    if detected == ".xlsx" and filename_extension != ".xlsx":
        raise UploadError("mismatched_content_type", "XLSX content must use the .xlsx extension.")
    if detected != ".xlsx" and data.startswith(b"PK\x03\x04"):
        raise UploadError("unsupported_type", "Archive formats other than XLSX are not supported.")
    return detected


def _extract_pdf(data: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise UploadError("extractor_unavailable", "PDF extraction requires pypdf.") from exc
    try:
        reader = PdfReader(io.BytesIO(data))
    except Exception as exc:
        raise UploadError("malformed_pdf", "PDF could not be read.") from exc
    if len(reader.pages) > _MAX_PDF_PAGES:
        raise UploadError("pdf_too_large", f"PDF exceeds the {_MAX_PDF_PAGES} page limit.")
    return "\n\n".join(page.extract_text() or "" for page in reader.pages)


def _extract_xlsx(data: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            members = archive.infolist()
            if len(members) > _MAX_XLSX_MEMBERS:
                raise UploadError("xlsx_too_many_members", "Workbook has too many ZIP members.")
            uncompressed = sum(member.file_size for member in members)
            if uncompressed > _MAX_XLSX_UNCOMPRESSED:
                raise UploadError("xlsx_too_large", "Workbook decompressed size exceeds the limit.")
            for member in members:
                name = member.filename.replace("\\", "/")
                if name.startswith("/") or ".." in name.split("/"):
                    raise UploadError(
                        "invalid_xlsx_member", "Workbook contains an unsafe member path."
                    )
    except UploadError:
        raise
    except (zipfile.BadZipFile, OSError) as exc:
        raise UploadError("malformed_xlsx", "Workbook is not a valid XLSX container.") from exc
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise UploadError("extractor_unavailable", "XLSX extraction requires openpyxl.") from exc
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise UploadError("malformed_xlsx", "Workbook could not be opened.") from exc
    parts: list[str] = []
    cells = 0
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                for value in row:
                    if value is None:
                        continue
                    cells += 1
                    if cells > _MAX_XLSX_CELLS:
                        raise UploadError(
                            "xlsx_too_many_cells", "Workbook has too many populated cells."
                        )
                    parts.append(str(value))
    finally:
        workbook.close()
    return "\n".join(parts)


def _extract_html_xml(data: bytes) -> str:
    from .source_ingest import _VisibleTextParser

    parser = _VisibleTextParser()
    parser.feed(data.decode("utf-8", errors="replace"))
    return "\n".join(parser.parts)


def _extract_text_file(data: bytes, detected: str) -> str:
    text = data.decode("utf-8", errors="replace")
    if detected == ".json":
        try:
            text = json.dumps(json.loads(text), ensure_ascii=False, indent=2)
        except json.JSONDecodeError:
            pass
    return text


def extract_text(filename: str, detected: str, data: bytes) -> str:
    """Extract bounded plain text through format-specific, non-executing adapters."""
    del filename
    if detected == ".pdf":
        text = _extract_pdf(data)
    elif detected == ".xlsx":
        text = _extract_xlsx(data)
    elif detected in {".html", ".xml"}:
        text = _extract_html_xml(data)
    elif detected in {".txt", ".csv", ".json"}:
        text = _extract_text_file(data, detected)
    else:
        raise UploadError("unsupported_type", f"Unsupported detected type {detected}.")
    if not text.strip():
        raise UploadError("empty_extraction", "The file produced no readable text.")
    if "\x00" in text:
        text = text.replace("\x00", " ")
    printable = sum(1 for char in text if char.isprintable() or char in "\n\t")
    if printable / max(1, len(text)) < _PRINTABLE_THRESHOLD:
        raise UploadError("non_text_content", "The file is not predominantly readable text.")
    return text


class UploadStorage:
    """Private on-disk storage for uploaded raw bytes with exclusive creation."""

    def __init__(self, root: Path, *, max_storage_bytes: int) -> None:
        self.root = root.expanduser().resolve()
        self.root.mkdir(parents=True, exist_ok=True)
        try:
            self.root.chmod(0o700)
        except OSError:
            pass
        self.max_storage_bytes = max_storage_bytes

    def _path(self, storage_key: str) -> Path:
        if not re.fullmatch(r"[0-9a-f]{32}\.bin", storage_key):
            raise UploadError("invalid_storage_key", "Storage key is invalid.")
        return self.root / storage_key

    def write(self, storage_key: str, data: bytes) -> None:
        target = self._path(storage_key)
        try:
            fd = os.open(target, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
        except FileExistsError as exc:
            raise UploadError("storage_conflict", "Storage key already exists.") from exc
        try:
            with os.fdopen(fd, "wb") as handle:
                handle.write(data)
                handle.flush()
                os.fsync(handle.fileno())
        except BaseException:
            try:
                target.unlink()
            except OSError:
                pass
            raise

    def read(self, storage_key: str) -> bytes:
        return self._path(storage_key).read_bytes()

    def delete(self, storage_key: str) -> None:
        try:
            self._path(storage_key).unlink()
        except FileNotFoundError:
            pass

    def total_bytes(self) -> int:
        return sum(path.stat().st_size for path in self.root.glob("*.bin") if path.is_file())

    def reserve(self, incoming_bytes: int) -> None:
        if self.total_bytes() + incoming_bytes > self.max_storage_bytes:
            raise UploadError("storage_quota_exceeded", "Upload storage quota would be exceeded.")


class UploadIngestor:
    """Validates, stores, and extracts user uploads into the review-only KB path."""

    def __init__(
        self,
        database: KnowledgeDatabase,
        storage: UploadStorage,
        *,
        max_file_bytes: int,
        audit: Callable[[str, Mapping[str, Any]], str] | None = None,
    ) -> None:
        self.database = database
        self.storage = storage
        self.max_file_bytes = max_file_bytes
        self.audit = audit or (lambda _event, _payload: "")

    def _audit(self, event_type: str, payload: Mapping[str, Any]) -> None:
        try:
            self.audit(event_type, dict(payload))
        except Exception:  # audit must never break the request path.
            pass

    @staticmethod
    def _metadata(fields: Mapping[str, str]) -> dict[str, Any]:
        metadata: dict[str, Any] = {}
        for key in _METADATA_FIELDS:
            if key in fields and fields[key]:
                metadata[key] = fields[key]
        families = metadata.get("materialFamilies")
        if isinstance(families, str):
            metadata["materialFamilies"] = [
                item.strip().casefold() for item in families.split(",") if item.strip()
            ]
        return metadata

    def ingest(
        self,
        part: MultipartPart,
        *,
        fields: Mapping[str, str] | None = None,
    ) -> UploadResult:
        metadata = self._metadata(fields or {})
        if len(part.data) > self.max_file_bytes:
            self._audit("kb.upload.rejected", {"filename": part.filename, "code": "file_too_large"})
            return UploadResult(
                "",
                part.filename,
                "failed",
                "",
                error_code="file_too_large",
                error_detail=f"File exceeds the {self.max_file_bytes} byte limit.",
            )
        raw_sha256 = hashlib.sha256(part.data).hexdigest()
        try:
            detected = _detect_extension(part.filename, part.data, part.declared_content_type)
        except UploadError as exc:
            self._audit("kb.upload.rejected", {"filename": part.filename, "code": exc.code})
            return UploadResult(
                "", part.filename, "failed", "", error_code=exc.code, error_detail=str(exc)
            )

        duplicate = self.database.find_upload_by_raw_sha256(raw_sha256)
        if duplicate is not None:
            self._audit(
                "kb.upload.accepted",
                {"filename": part.filename, "rawSha256": raw_sha256, "duplicate": True},
            )
            return UploadResult(
                str(duplicate["uploadId"]),
                part.filename,
                "duplicate",
                "",
                duplicate_of=str(duplicate["uploadId"]),
            )

        self.storage.reserve(len(part.data))
        storage_key = f"{uuid.uuid4().hex}.bin"
        upload_id = f"upload-{uuid.uuid4()}"
        try:
            self.storage.write(storage_key, part.data)
        except UploadError:
            self._audit("kb.upload.rejected", {"filename": part.filename, "code": "storage_failed"})
            return UploadResult(
                "",
                part.filename,
                "failed",
                "",
                error_code="storage_failed",
                error_detail="The file could not be stored.",
            )
        self.database.create_upload(
            upload_id,
            original_filename=part.filename,
            storage_key=storage_key,
            raw_sha256=raw_sha256,
            size_bytes=len(part.data),
            declared_content_type=part.declared_content_type,
            detected_content_type=detected,
            extension=detected,
            status="validated",
        )
        job_id = self.database.create_job(
            "extract",
            upload_id,
            idempotency_key=raw_sha256,
            step="queued",
            progress_total=1,
            detail=json.dumps(metadata, ensure_ascii=False, sort_keys=True),
        )
        self._audit(
            "kb.upload.accepted",
            {
                "uploadId": upload_id,
                "filename": part.filename,
                "rawSha256": raw_sha256,
                "sizeBytes": len(part.data),
                "detectedContentType": detected,
            },
        )
        return UploadResult(upload_id, part.filename, "validated", job_id)

    def extract(self, upload_id: str) -> dict[str, Any]:
        """Extract a stored upload, register a review source/document, and link the upload."""
        upload = self.database.get_upload(upload_id)
        if upload is None or upload["status"] != "validated":
            raise UploadError("invalid_upload_state", "Upload is not ready for extraction.")
        metadata: dict[str, Any] = {}
        extract_job = self.database.find_job("extract", upload_id)
        if extract_job and extract_job.get("detail"):
            try:
                parsed = json.loads(extract_job["detail"])
                if isinstance(parsed, dict):
                    metadata = parsed
            except json.JSONDecodeError:
                metadata = {}
        data = self.storage.read(upload["storageKey"])
        try:
            text = extract_text(upload["originalFilename"], upload["extension"], data)
        except UploadError as exc:
            self.database.update_upload(
                upload_id, status="failed", error_code=exc.code, error_detail=str(exc)
            )
            self._audit("kb.upload.rejected", {"uploadId": upload_id, "code": exc.code})
            raise
        families = tuple(
            sorted({str(item).casefold() for item in metadata.get("materialFamilies", [])})
        )
        source_id = f"upload-{upload['rawSha256'][:16]}"
        title = str(metadata.get("title") or f"User upload: {upload['originalFilename']}")[:300]
        publisher = str(metadata.get("publisher") or "not provided")[:200]
        authority_tier = str(metadata.get("authorityTier") or "secondary")
        scale = str(metadata.get("scale") or "")
        license_note = "user-supplied; reuse unknown"
        limitations = "not independently verified"
        self.database.register_user_upload_source(
            source_id=source_id,
            title=title,
            publisher=publisher,
            source_type="USER_UPLOAD",
            material_families=families,
            jurisdiction=str(metadata.get("jurisdiction") or ""),
            license_note=license_note,
            limitations=limitations,
            authority_tier=authority_tier,
            scale=scale,
            metadata_origin="operator-supplied" if metadata else "system-default",
        )
        result = self.database.ingest_upload_document(
            source_id=source_id,
            title=title,
            material_families=families,
            content=text,
            content_type=upload["detectedContentType"],
        )
        self.database.update_upload(
            upload_id,
            status="ingested",
            source_id=source_id,
            document_id=result.document_id,
            error_code=None,
            error_detail=None,
        )
        self._audit(
            "kb.document.extracted",
            {
                "uploadId": upload_id,
                "sourceId": source_id,
                "documentId": result.document_id,
                "chunkCount": result.chunk_count,
                "contentSha256": result.content_sha256,
            },
        )
        return {
            "uploadId": upload_id,
            "sourceId": source_id,
            "documentId": result.document_id,
            "chunkCount": result.chunk_count,
        }
